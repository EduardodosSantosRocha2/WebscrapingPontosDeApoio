import os
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException

import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill


orangeFill = PatternFill(start_color='FFC000',end_color='FFC000',fill_type='solid')

blueFill = PatternFill(start_color='00B0F0',end_color='00B0F0',fill_type='solid')

greenFill = PatternFill(start_color='00c04b',end_color='00c04b',fill_type='solid')

purpleFill = PatternFill(start_color='3139E4',end_color='3139E4',fill_type='solid')

yeloyFill = PatternFill(start_color='f7d917',end_color='f7d917', fill_type='solid')



wb = Workbook()

del wb['Sheet']

sheet = wb.create_sheet('Nome do Sheet')

sheet.cell(1,1).value = "E-mail"
sheet.cell(1,1).fill = orangeFill

sheet['B1'] = "Endereço"
sheet['B1'].fill = blueFill

sheet['C1'] = "Mantenedora"
sheet['C1'].fill = greenFill


sheet['D1'] = "CNPJ da mantenedora"
sheet['D1'].fill = purpleFill


service = Service()
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)


url = 'https://cvv.org.br/'
driver.get(url)


wait = WebDriverWait(driver, 10)


botao_aceitartodos = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, 'aceitar-todos')))
botao_aceitartodos.click()

def busca(estado, cidade):
    estado_select = wait.until(EC.presence_of_element_located((By.ID, 'estado')))
    estado_dropdown = Select(estado_select)
    estado_dropdown.select_by_value(estado)

    cidade_select = wait.until(EC.presence_of_element_located((By.ID, 'cidade')))
    cidade_dropdown = Select(cidade_select)
    cidade_dropdown.select_by_value(cidade)


    buscar_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[type="submit"]')))
    buscar_button.click()


    time.sleep(2) 




    elementos_p = wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, 'p')))


    elementos_p = elementos_p[34:38]

    return elementos_p



def gerarPlanilha(elementosDicionario, nomePlanilha):
    # Verifica se o arquivo já existe
    if os.path.exists(f"{nomePlanilha}.xlsx"):
        df = pd.read_excel(f"{nomePlanilha}.xlsx", engine='openpyxl')
        countLines = df.shape[0] + 2  # Considera a última linha com dados e o cabeçalho
    else:
        countLines = 2  # Começa a partir da segunda linha (a primeira é o cabeçalho)

    for j, elementos in enumerate(elementosDicionario):
        for i, valores in enumerate(elementos):
            aux = valores.split(":")
            if len(aux) > 1:  # Verifica se a divisão foi bem-sucedida
                sheet.cell(countLines + j, i + 1).value = aux[1].strip()  # Adiciona valor à célula

    wb.save(f"{nomePlanilha}.xlsx")





def extrairTexto(elementosDicionario, estado, cidade):    
    elementos_cidades = busca(estado, cidade)
    elementos_text = [elemento.text for elemento in elementos_cidades]
    print(f"Elementos Uberlândia: {elementos_text}")
    elementosDicionario.append(elementos_text)



elementosDicionario = []
extrairTexto(elementosDicionario, 'MG', 'BELO HORIZONTE')
extrairTexto(elementosDicionario, 'MG', 'IPATINGA')
extrairTexto(elementosDicionario, 'MG', 'ITABIRA')
extrairTexto(elementosDicionario, 'MG', 'MONTES CLAROS')
extrairTexto(elementosDicionario, 'MG', 'UBERABA')
extrairTexto(elementosDicionario, 'MG', 'UBERLÂNDIA')
gerarPlanilha(elementosDicionario, "CVV")

time.sleep(1)

# Feche o navegador
driver.quit()
